"""
Système d'import Excel pour GestMat.
Supporte un fichier multi-feuilles avec une feuille par entité.
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime, date as dt_date

from django.db import transaction

from .models import (
    Depot, Unite, TypeOperation,
    ComptePrincipale, SousCompte,
    Fournisseur, Beneficiaire,
    Produit, SocieteGCS,
    MembreCommission, MembreCommissionReforme,
    AnneeExercice, MatieresDepot,
)


# ─────────────────────────── UTILITAIRES ───────────────────────────

def _sheet_to_rows(sheet):
    """Retourne une liste de dict {header_normalisé: valeur} pour chaque ligne."""
    data = list(sheet.iter_rows(values_only=True))
    if not data:
        return []
    headers = [_normalize(h) for h in data[0]]
    result = []
    for row in data[1:]:
        if all(v is None for v in row):
            continue
        result.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})
    return result


def _normalize(s):
    """Normalise un en-tête : minuscules, sans accents, sans espaces superflus."""
    import unicodedata
    if s is None:
        return ''
    s = str(s).lower().strip()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    for old, new in [(' ', '_'), ('°', ''), ('.', ''), ('/', '_'), ('-', '_'), ("'", '_')]:
        s = s.replace(old, new)
    return s


def _get(row, *keys, default=''):
    """Récupère la première valeur non-nulle parmi les clés (normalisées)."""
    for k in keys:
        nk = _normalize(k)
        if nk in row and row[nk] is not None:
            v = row[nk]
            return str(v).strip() if not isinstance(v, (int, float)) else v
    return default


def _str(row, *keys):
    v = _get(row, *keys, default='')
    return str(v).strip() if v != '' else ''


def _int(row, *keys, default=0):
    v = _get(row, *keys, default=None)
    if v is None or v == '':
        return default
    try:
        return int(float(str(v)))
    except (ValueError, TypeError):
        return default


def _float(row, *keys, default=None):
    v = _get(row, *keys, default=None)
    if v is None or v == '':
        return default
    try:
        return float(str(v))
    except (ValueError, TypeError):
        return default


def _date(v):
    """Parse une valeur en date Python."""
    if v is None or v == '':
        return None
    if isinstance(v, dt_date):
        return v
    if hasattr(v, 'date'):
        return v.date()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d.%m.%Y'):
        try:
            return datetime.strptime(str(v).strip(), fmt).date()
        except ValueError:
            continue
    return None


# ─────────────────────────── IMPORTEURS PAR MODÈLE ───────────────────────────

def _import_depots(rows):
    created = updated = 0
    errors = []
    for i, row in enumerate(rows, 2):
        code = _str(row, 'Code', 'code_depot', 'depot', 'code depot')
        if not code:
            errors.append(f"Ligne {i} : Code dépôt manquant")
            continue
        _, is_new = Depot.objects.get_or_create(code=code)
        created += is_new
        updated += not is_new
    return created, updated, errors


def _import_unites(rows):
    created = updated = 0
    errors = []
    for i, row in enumerate(rows, 2):
        lib = _str(row, 'Libelle', 'libellé', 'unite', 'unité', 'designation', 'désignation')
        if not lib:
            errors.append(f"Ligne {i} : Libellé manquant")
            continue
        _, is_new = Unite.objects.get_or_create(libelle=lib)
        created += is_new
        updated += not is_new
    return created, updated, errors


def _import_types_operation(rows):
    created = updated = 0
    errors = []
    for i, row in enumerate(rows, 2):
        lib = _str(row, 'Libelle', 'libellé', 'type', 'type_operation', "type d'opération", "type d'operation")
        if not lib:
            errors.append(f"Ligne {i} : Libellé manquant")
            continue
        _, is_new = TypeOperation.objects.get_or_create(libelle=lib)
        created += is_new
        updated += not is_new
    return created, updated, errors


def _import_comptes_principaux(rows):
    created = updated = 0
    errors = []
    for i, row in enumerate(rows, 2):
        num = _get(row, 'N°Compte', 'num_compte', 'numero', 'n° compte', 'numéro compte', 'compte', default=None)
        famille = _str(row, 'Famille', 'libelle', 'libellé', 'designation', 'désignation')
        if num is None or str(num).strip() == '':
            errors.append(f"Ligne {i} : N° compte manquant")
            continue
        if not famille:
            errors.append(f"Ligne {i} : Famille manquante")
            continue
        try:
            num = int(float(str(num)))
        except (ValueError, TypeError):
            errors.append(f"Ligne {i} : N° compte invalide ({num})")
            continue
        obj, is_new = ComptePrincipale.objects.get_or_create(num_compte=num, defaults={'famille': famille})
        if not is_new and obj.famille != famille:
            obj.famille = famille
            obj.save()
        created += is_new
        updated += not is_new
    return created, updated, errors


def _import_sous_comptes(rows):
    created = updated = 0
    errors = []
    for i, row in enumerate(rows, 2):
        num_cp_raw = _get(row, 'N°CompteP', 'num_compte', 'compte_principal', 'n° compte', 'compte', default=None)
        num_sc_raw = _get(row, 'N°SousCompte', 'num_sous_compte', 'sous_compte', 'n° sous compte', default=None)
        famille = _str(row, 'Famille', 'famille_sc', 'libelle', 'libellé')
        if num_cp_raw is None or num_sc_raw is None:
            errors.append(f"Ligne {i} : Numéros manquants")
            continue
        try:
            num_cp = int(float(str(num_cp_raw)))
            num_sc = float(str(num_sc_raw))
        except (ValueError, TypeError):
            errors.append(f"Ligne {i} : Numéros invalides ({num_cp_raw}, {num_sc_raw})")
            continue
        try:
            cp = ComptePrincipale.objects.get(num_compte=num_cp)
        except ComptePrincipale.DoesNotExist:
            errors.append(f"Ligne {i} : Compte principal {num_cp} introuvable (importez ComptesPrincipaux d'abord)")
            continue
        obj, is_new = SousCompte.objects.get_or_create(
            compte_principal=cp, num_sous_compte=num_sc,
            defaults={'famille_sc': famille}
        )
        if not is_new and famille:
            obj.famille_sc = famille
            obj.save()
        created += is_new
        updated += not is_new
    return created, updated, errors


def _import_fournisseurs(rows):
    created = updated = 0
    errors = []
    for i, row in enumerate(rows, 2):
        nom = _str(row, 'Nom', 'fournisseur', 'raison sociale', 'raison_sociale', 'societe', 'société')
        if not nom:
            errors.append(f"Ligne {i} : Nom manquant")
            continue
        defaults = {
            'adresse':     _str(row, 'Adresse', 'adresse'),
            'telephone':   _str(row, 'Telephone', 'téléphone', 'tel', 'tél', 'phone'),
            'fax':         _str(row, 'Fax', 'fax'),
            'code_postal': _str(row, 'CodePostal', 'code_postal', 'code postal', 'cp'),
            'ville':       _str(row, 'Ville', 'ville'),
            'region':      _str(row, 'Region', 'région', 'wilaya'),
            'pays':        _str(row, 'Pays', 'pays'),
            'fonction':    _str(row, 'Fonction', 'fonction', 'activite', 'activité'),
        }
        obj, is_new = Fournisseur.objects.get_or_create(nom=nom, defaults=defaults)
        if not is_new:
            changed = False
            for k, v in defaults.items():
                if v and getattr(obj, k) != v:
                    setattr(obj, k, v)
                    changed = True
            if changed:
                obj.save()
        created += is_new
        updated += not is_new
    return created, updated, errors


def _import_beneficiaires(rows):
    created = updated = 0
    errors = []
    for i, row in enumerate(rows, 2):
        nom = _str(row, 'Nom', 'beneficiaire', 'bénéficiaire', 'service', 'structure')
        if not nom:
            errors.append(f"Ligne {i} : Nom manquant")
            continue
        responsable = _str(row, 'Responsable', 'responsable', 'chef', 'chef_service', 'chef service')
        obj, is_new = Beneficiaire.objects.get_or_create(nom=nom, defaults={'responsable': responsable})
        if not is_new and responsable:
            obj.responsable = responsable
            obj.save()
        created += is_new
        updated += not is_new
    return created, updated, errors


def _import_produits(rows):
    created = updated = 0
    errors = []
    for i, row in enumerate(rows, 2):
        nomenclature = _str(row, 'Nomenclature', 'code', 'reference', 'référence', 'ref', 'réf', 'n°article', 'code article')
        designation  = _str(row, 'Designation', 'désignation', 'libelle', 'libellé', 'article', 'produit', 'intitule', 'intitulé')
        if not nomenclature:
            errors.append(f"Ligne {i} : Nomenclature manquante")
            continue
        if not designation:
            errors.append(f"Ligne {i} : Désignation manquante")
            continue

        # Unité — auto-create
        unite_lib = _str(row, 'Unite', 'unité', 'um', 'u.m.', 'UM')
        unite = None
        if unite_lib:
            unite, _ = Unite.objects.get_or_create(libelle=unite_lib)

        # Compte principal
        cp = None
        num_cp = _get(row, 'N°Compte', 'num_compte', 'compte', 'n° compte', default=None)
        if num_cp is not None and str(num_cp).strip():
            try:
                cp = ComptePrincipale.objects.get(num_compte=int(float(str(num_cp))))
            except (ComptePrincipale.DoesNotExist, ValueError):
                pass

        # Sous-compte
        sc = None
        num_sc = _get(row, 'N°SousCompte', 'num_sous_compte', 'sous_compte', 'n° sous compte', default=None)
        if num_sc is not None and cp:
            try:
                sc = SousCompte.objects.get(compte_principal=cp, num_sous_compte=float(str(num_sc)))
            except (SousCompte.DoesNotExist, ValueError):
                pass

        stock = _int(row, 'StockGlobal', 'stock_global', 'stock global', 'stock', 'qte', 'quantite', 'quantité', default=0)
        spec  = _str(row, 'Specification', 'spécification', 'spec', 'specification')
        obs   = _str(row, 'Observation', 'obs', 'remarque')

        defaults = {
            'designation':    designation,
            'unite':          unite,
            'compte_principal': cp,
            'sous_compte':    sc,
            'stock_global':   stock,
            'specification':  spec,
            'observation':    obs,
        }
        obj, is_new = Produit.objects.get_or_create(nomenclature=nomenclature, defaults=defaults)
        if not is_new:
            for k, v in defaults.items():
                if v is not None:
                    setattr(obj, k, v)
            obj.save()
        created += is_new
        updated += not is_new
    return created, updated, errors


def _import_membres_commission(rows):
    created = updated = 0
    errors = []
    for i, row in enumerate(rows, 2):
        nom     = _str(row, 'Nom', 'membre', 'nom_prenom', 'nom prénom')
        qualite = _str(row, 'Qualite', 'qualité', 'fonction', 'grade', 'titre')
        if not nom:
            errors.append(f"Ligne {i} : Nom manquant")
            continue
        obj, is_new = MembreCommission.objects.get_or_create(nom=nom, defaults={'qualite': qualite})
        if not is_new and qualite:
            obj.qualite = qualite
            obj.save()
        created += is_new
        updated += not is_new
    return created, updated, errors


def _import_membres_reforme(rows):
    created = updated = 0
    errors = []
    for i, row in enumerate(rows, 2):
        nom     = _str(row, 'Nom', 'membre')
        qualite = _str(row, 'Qualite', 'qualité', 'fonction', 'grade')
        if not nom:
            errors.append(f"Ligne {i} : Nom manquant")
            continue
        obj, is_new = MembreCommissionReforme.objects.get_or_create(nom=nom, defaults={'qualite': qualite})
        if not is_new and qualite:
            obj.qualite = qualite
            obj.save()
        created += is_new
        updated += not is_new
    return created, updated, errors


def _import_annees_exercice(rows):
    created = updated = 0
    errors = []
    for i, row in enumerate(rows, 2):
        annee_raw  = _get(row, 'Annee', 'année', 'exercice', default=None)
        debut_raw  = _get(row, 'DateDebut', 'date_debut', 'date début', 'debut', 'début', default=None)
        fin_raw    = _get(row, 'DateFin', 'date_fin', 'date fin', 'fin', default=None)
        if annee_raw is None or str(annee_raw).strip() == '':
            errors.append(f"Ligne {i} : Année manquante")
            continue
        try:
            annee = int(float(str(annee_raw)))
        except (ValueError, TypeError):
            errors.append(f"Ligne {i} : Année invalide ({annee_raw})")
            continue

        # Support dates Excel (datetime objects) ou strings
        dd = _date(debut_raw) or dt_date(annee, 1, 1)
        df = _date(fin_raw)   or dt_date(annee, 12, 31)

        obj, is_new = AnneeExercice.objects.get_or_create(
            annee=annee,
            defaults={'date_debut': dd, 'date_fin': df}
        )
        if not is_new:
            obj.date_debut = dd
            obj.date_fin   = df
            obj.save()
        created += is_new
        updated += not is_new
    return created, updated, errors


# ─────────────────────────── REGISTRE ───────────────────────────

# (clé interne, label affiché, nom feuille Excel, fonction import)
IMPORTERS = {
    'depots':             ("Dépôts",               "Depots",             _import_depots),
    'unites':             ("Unités",                "Unites",             _import_unites),
    'types_operation':    ("Types d'opération",     "TypesOperation",     _import_types_operation),
    'comptes_principaux': ("Comptes principaux",    "ComptesPrincipaux",  _import_comptes_principaux),
    'sous_comptes':       ("Sous-comptes",           "SousComptes",        _import_sous_comptes),
    'fournisseurs':       ("Fournisseurs",            "Fournisseurs",       _import_fournisseurs),
    'beneficiaires':      ("Bénéficiaires",           "Beneficiaires",      _import_beneficiaires),
    'produits':           ("Produits / Matières",    "Produits",           _import_produits),
    'membres_commission': ("Membres commission",     "MembresCommission",  _import_membres_commission),
    'membres_reforme':    ("Membres réforme",        "MembresReforme",     _import_membres_reforme),
    'annees_exercice':    ("Années d'exercice",      "AnneesExercice",     _import_annees_exercice),
}

# Index inversé : sheet_name_normalisé → clé
_SHEET_INDEX = {}
_ALIASES = {
    'depot': 'depots', 'dépôt': 'depots', 'dépôts': 'depots',
    'unite': 'unites', 'unité': 'unites', 'unités': 'unites',
    'typeoperation': 'types_operation', 'typeoperations': 'types_operation',
    'compteprincipal': 'comptes_principaux', 'comptes': 'comptes_principaux',
    'souscompte': 'sous_comptes', 'sous_compte': 'sous_comptes',
    'fournisseur': 'fournisseurs',
    'beneficiaire': 'beneficiaires', 'bénéficiaire': 'beneficiaires', 'bénéficiaires': 'beneficiaires',
    'produit': 'produits', 'matieres': 'produits', 'matières': 'produits',
    'membrecommission': 'membres_commission', 'commission': 'membres_commission',
    'membrereforme': 'membres_reforme', 'reforme': 'membres_reforme', 'réforme': 'membres_reforme',
    'annee': 'annees_exercice', 'exercice': 'annees_exercice', 'annees': 'annees_exercice',
}
for key, (_, sheet_name, _fn) in IMPORTERS.items():
    _SHEET_INDEX[_normalize(sheet_name)] = key
_SHEET_INDEX.update({_normalize(k): v for k, v in _ALIASES.items()})


# ─────────────────────────── PROCESS FICHIER ───────────────────────────

def process_excel(file_obj):
    """
    Traite un fichier Excel uploadé.
    Retourne une liste de résultats par feuille :
    [{'key', 'label', 'sheet', 'rows', 'created', 'updated', 'errors', 'skipped'}]
    """
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    results = []

    for sheet_name in wb.sheetnames:
        key = _SHEET_INDEX.get(_normalize(sheet_name))
        if key is None:
            results.append({
                'key': None, 'label': sheet_name, 'sheet': sheet_name,
                'rows': 0, 'created': 0, 'updated': 0,
                'errors': [], 'skipped': True,
                'note': f'Feuille non reconnue — ignorée',
            })
            continue

        label, _, fn = IMPORTERS[key]
        sheet = wb[sheet_name]
        rows  = _sheet_to_rows(sheet)

        if not rows:
            results.append({
                'key': key, 'label': label, 'sheet': sheet_name,
                'rows': 0, 'created': 0, 'updated': 0,
                'errors': ['Feuille vide'], 'skipped': False,
            })
            continue

        try:
            with transaction.atomic():
                created, updated, errors = fn(rows)
        except Exception as exc:
            results.append({
                'key': key, 'label': label, 'sheet': sheet_name,
                'rows': len(rows), 'created': 0, 'updated': 0,
                'errors': [f"Erreur inattendue : {exc}"], 'skipped': False,
            })
            continue

        results.append({
            'key': key, 'label': label, 'sheet': sheet_name,
            'rows': len(rows), 'created': created, 'updated': updated,
            'errors': errors, 'skipped': False,
        })

    wb.close()
    return results


# ─────────────────────────── GÉNÉRATION TEMPLATE ───────────────────────────

def generate_template():
    """Génère un fichier Excel template multi-feuilles."""
    wb = Workbook()
    wb.remove(wb.active)

    H_FILL  = PatternFill("solid", fgColor="1E3A8A")
    H_FONT  = Font(bold=True, color="FFFFFF", size=10, name='Calibri')
    EX_FILL = PatternFill("solid", fgColor="EFF6FF")
    EX_FONT = Font(size=10, name='Calibri', color="1E40AF", italic=True)
    border  = Border(
        left=Side(style='thin', color='BFDBFE'),
        right=Side(style='thin', color='BFDBFE'),
        top=Side(style='thin', color='BFDBFE'),
        bottom=Side(style='thin', color='BFDBFE'),
    )

    def sheet(name, headers, examples, notes=None):
        ws = wb.create_sheet(name)
        ws.sheet_properties.tabColor = "1E40AF"

        # En-têtes
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.font   = H_FONT
            cell.fill   = H_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            ws.column_dimensions[get_column_letter(c)].width = max(len(str(h)) + 6, 18)
        ws.row_dimensions[1].height = 22

        # Exemples
        for r, ex_row in enumerate(examples, 2):
            for c, v in enumerate(ex_row, 1):
                cell = ws.cell(r, c, v)
                cell.font   = EX_FONT
                cell.fill   = EX_FILL
                cell.border = border
                cell.alignment = Alignment(vertical='center')

        ws.freeze_panes = 'A2'

        # Note optionnelle dans colonne après les headers
        if notes:
            note_col = len(headers) + 2
            ws.cell(1, note_col, "📌 Notes").font = Font(bold=True, color="92400E", size=9)
            for r, note in enumerate(notes, 2):
                c = ws.cell(r, note_col, note)
                c.font = Font(color="92400E", size=9, italic=True)
        return ws

    sheet("Depots",
          ["Code"],
          [["D01"], ["D02"], ["D03"]],
          ["Code unique du dépôt", "", ""])

    sheet("Unites",
          ["Libelle"],
          [["Boite"], ["Unité"], ["Kg"], ["Litre"], ["Paire"], ["Pièce"], ["Flacon"]])

    sheet("TypesOperation",
          ["Libelle"],
          [["Achat direct"], ["Marché"], ["Don"], ["Balance d'ouverture"]])

    sheet("ComptesPrincipaux",
          ["N°Compte", "Famille"],
          [[31, "Matières premières et fournitures"],
           [32, "Matières consommables"],
           [33, "Produits en cours"]],
          ["Entier unique", "Intitulé du compte", ""])

    sheet("SousComptes",
          ["N°CompteP", "N°SousCompte", "Famille"],
          [[31, 31.1, "Médicaments"],
           [31, 31.2, "Dispositifs médicaux"],
           [32, 32.1, "Produits chimiques"]],
          ["N° compte principal existant", "Nombre décimal", "Intitulé"])

    sheet("Fournisseurs",
          ["Nom", "Adresse", "Telephone", "Fax", "Ville", "Region", "Pays"],
          [["SARL Pharma", "12 rue Didouche", "021 00 00 00", "021 00 00 01", "Alger", "Alger", "Algérie"],
           ["ETS Médical", "Zone ind. Rouiba", "023 00 00 00", "", "Rouiba", "Boumerdès", "Algérie"]])

    sheet("Beneficiaires",
          ["Nom", "Responsable"],
          [["Service Pharmacie", "Dr. Kaci Amar"],
           ["Bloc opératoire", "Dr. Bouazza Fatima"],
           ["Service urgences", ""],
           ["Médecine interne", "Dr. Hamdi Said"]])

    sheet("Produits",
          ["Nomenclature", "Designation", "Unite", "N°Compte", "N°SousCompte", "StockGlobal", "Specification"],
          [["MED001", "Paracetamol 500mg cp", "Boite", 31, 31.1, 0, "500mg — 16 comprimés/boite"],
           ["MED002", "Amoxicilline 500mg", "Boite", 31, 31.1, 0, "500mg gélules"],
           ["MAT001", "Seringue 10ml", "Pièce", 32, 32.1, 0, "Jetable stérile"],
           ["MAT002", "Gants stériles L", "Paire", 32, 32.1, 0, "Taille L"]],
          ["Code unique", "Désignation complète", "Doit exister (auto-créé)", "Optionnel", "Optionnel", "Quantité initiale", ""])

    sheet("MembresCommission",
          ["Nom", "Qualite"],
          [["M. Martin Ahmed", "Directeur Général"],
           ["Mme. Benali Fatima", "Chef de service pharmacie"],
           ["M. Khelil Rachid", "Technicien supérieur"]])

    sheet("MembresReforme",
          ["Nom", "Qualite"],
          [["M. Boudiaf Karim", "Ingénieur biomédical"],
           ["Mme. Saidi Amina", "Cadre administratif"]])

    sheet("AnneesExercice",
          ["Annee", "DateDebut", "DateFin"],
          [[2023, "01/01/2023", "31/12/2023"],
           [2024, "01/01/2024", "31/12/2024"],
           [2025, "01/01/2025", "31/12/2025"]],
          ["Entier (ex: 2024)", "Format jj/mm/aaaa", "Format jj/mm/aaaa"])

    # Feuille README
    readme = wb.create_sheet("📋 README", 0)
    readme.sheet_properties.tabColor = "047857"
    readme.column_dimensions['A'].width = 60
    readme.column_dimensions['B'].width = 40
    instructions = [
        ("GUIDE D'UTILISATION — GestMat Import Excel", ""),
        ("", ""),
        ("1. Ne modifiez pas les noms des feuilles (onglets)", ""),
        ("2. Ne modifiez pas la ligne d'en-tête (ligne 1)", ""),
        ("3. Supprimez les lignes d'exemple avant d'importer vos données", ""),
        ("4. Les colonnes en italique bleu sont des exemples à remplacer", ""),
        ("5. Importez dans l'ordre suivant pour respecter les dépendances :", ""),
        ("   → Depots, Unites, TypesOperation", ""),
        ("   → ComptesPrincipaux", ""),
        ("   → SousComptes (nécessite ComptesPrincipaux)", ""),
        ("   → Fournisseurs, Beneficiaires", ""),
        ("   → Produits (nécessite Unites + Comptes si liés)", ""),
        ("   → MembresCommission, MembresReforme, AnneesExercice", ""),
        ("", ""),
        ("6. Vous pouvez importer une seule feuille ou toutes à la fois", ""),
        ("7. Les doublons sont mis à jour (pas dupliqués)", ""),
        ("8. En cas d'erreur, seule la ligne fautive est ignorée", ""),
    ]
    for r, (a, b) in enumerate(instructions, 1):
        ca = readme.cell(r, 1, a)
        if r == 1:
            ca.font = Font(bold=True, size=13, color="047857")
        elif a.startswith("   →"):
            ca.font = Font(size=10, color="1D4ED8")
        elif a and a[0].isdigit():
            ca.font = Font(bold=True, size=10, color="1F2937")
        else:
            ca.font = Font(size=10, color="6B7280")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
